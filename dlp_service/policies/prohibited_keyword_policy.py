import zipfile
# import fitz
import pymupdf
import csv
from io import BytesIO
from docx import Document
from openpyxl import load_workbook

# valid_file_types = [".txt", ".doc", ".docx", ".csv", ".xlsx", ".pdf"]
prohibited_words = ["apple", "banana", "orange"]

def prohibited_keyword_SEARCH_for_TEXT(text):
    count = 0
    prohibited_word_encountered = False
    for word in prohibited_words:
        print(f"iteration: {count} | word: {word}")
        if word in text.lower():
            prohibited_word_encountered = True
            print(f"prohibited word found in text!!")
            if prohibited_word_encountered == True:
                return {
                    "allowed": False,
                    "reason": "prohibited word found in TXT/CSV file"
                }
        elif count < len(prohibited_words) - 1:
            print(f"iteration: {count} | word: {word} | current word is not found -> proceeding to next word")
        else:
            print(f"prohibited word NOT found in text!!")
            return {
                "allowed": True,
                "reason": "prohibited word NOT found in received file"
            }
        count += 1

def check_policy(metadata, file):
    print(f"metadata: {metadata}")
    global valid_file_types
    file.seek(0)
    data = file.read()
    # print(f"data: {data[: 10]} | it's type is: {type(data)}")
    print(f"data: {data} | it's type is: {type(data)}")
    if data.startswith(b"%PDF-"):
        print(f"file is PDF")
        file.seek(0)
        # doc = fitz.open(stream = data, filetype = "pdf")
        doc = pymupdf.open(stream = data, filetype = "pdf")
        text = ""
        for page in doc:
            text += page.get_text()
        print(f"pdf data is: {text} | its type is: {type(text)}")
        return prohibited_keyword_SEARCH_for_TEXT(text)
    elif zipfile.is_zipfile(file):
        file.seek(0)
        with zipfile.ZipFile(file) as z:
            names = z.namelist()
            if any(name.startswith("word/") for name in names):
                print(f"file is DOCX")
                doc = Document(BytesIO(data))
                text = ""
                for para in doc.paragraphs:
                    text += para.text
                print(f"the doc data is: {text} | it's type is: {type(text)}")
                file.seek(0)
                return prohibited_keyword_SEARCH_for_TEXT(text)
            elif any(name.startswith("xl/") for name in names):
                print(f"file is XLSX")
                workbook = load_workbook(BytesIO(data))
                text = ""
                for sheet in workbook.worksheets:
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                text += str(cell.value) + " "
                print(f"the exel data is: {text} | it's type is: {type(text)}")
                file.seek(0)
                return prohibited_keyword_SEARCH_for_TEXT(text)
    else:
        file.seek(0)
        try:
            text = data.decode("utf-8")
            print(f"type of text is: {type(text)}")
            print(f"text: {text}")
            print(f"prolly file is TXT or CSV")
            # if any(word.lower() in text.lower() for word.lower() in prohibited_words):
            return prohibited_keyword_SEARCH_for_TEXT(text)

        except UnicodeDecodeError:
            print(f"file is NOT TXT")
            print(f"file might be of non registered types")
            return {
                "allowed": True,
                "reason": "non applicable file"
            }